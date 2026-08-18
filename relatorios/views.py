import io
import unicodedata
from datetime import date, timedelta

from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDate, TruncMonth
from django.http import HttpResponse
from django.utils import timezone
from rest_framework import views
from rest_framework.permissions import AllowAny
from rest_framework.response import Response

from ifood.models import PedidoIFood, ItemPedidoIFood
from pdv.models import ItemPedidoPDV
from eventos.models import ItemEvento


class CsrfExemptMixin:
    authentication_classes = []


def _normalizar_nome(nome):
    """Chave de agrupamento sem acento/caixa — nenhum canal garante FK pra
    pdv.Produto (ItemPedidoIFood nunca tem), então o agrupamento entre canais
    é sempre por texto do nome do item, nunca por catálogo."""
    n = unicodedata.normalize('NFKD', nome or '').encode('ascii', 'ignore').decode('ascii')
    return ' '.join(n.strip().lower().split())


class RelatorioIFoodView(CsrfExemptMixin, views.APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        params = request.query_params
        formato = params.get('formato', 'json')
        agrupamento = params.get('agrupamento', 'dia')

        hoje = timezone.localtime(timezone.now()).date()
        try:
            data_inicio = date.fromisoformat(params['data_inicio']) if params.get('data_inicio') else hoje - timedelta(days=29)
        except ValueError:
            data_inicio = hoje - timedelta(days=29)
        try:
            data_fim = date.fromisoformat(params['data_fim']) if params.get('data_fim') else hoje
        except ValueError:
            data_fim = hoje

        if data_inicio > data_fim:
            data_inicio, data_fim = data_fim, data_inicio

        qs = PedidoIFood.objects.filter(
            ifood_criado_em__date__gte=data_inicio,
            ifood_criado_em__date__lte=data_fim,
        )

        resumo = self._calc_resumo(qs)
        agrupado = self._calc_agrupado(qs, agrupamento)

        dados = {
            'periodo': {'inicio': str(data_inicio), 'fim': str(data_fim)},
            'agrupamento': agrupamento,
            'resumo': resumo,
            'agrupado': agrupado,
        }

        if formato == 'excel':
            return self._export_excel(dados)
        if formato == 'pdf':
            return self._export_pdf(dados)

        return Response(dados)

    # ──────────────────────────────────────────────────────────────────────────

    def _calc_resumo(self, qs):
        agg = qs.aggregate(
            total=Count('id'),
            receita=Sum('total_valor'),
            cancelados=Count('id', filter=Q(status='CANCELLED')),
            delivery=Count('id', filter=Q(order_type='DELIVERY')),
            takeout=Count('id', filter=Q(order_type='TAKEOUT')),
            indoor=Count('id', filter=Q(order_type='INDOOR')),
        )
        total = agg['total'] or 0
        receita = float(agg['receita'] or 0)
        cancelados = agg['cancelados'] or 0
        nao_cancelados = total - cancelados
        ticket = receita / nao_cancelados if nao_cancelados else 0
        return {
            'total_pedidos': total,
            'receita_total': round(receita, 2),
            'ticket_medio': round(ticket, 2),
            'cancelados': cancelados,
            'taxa_cancelamento': round(cancelados / total * 100, 1) if total else 0,
            'delivery': agg['delivery'] or 0,
            'takeout': agg['takeout'] or 0,
            'indoor': agg['indoor'] or 0,
        }

    def _calc_agrupado(self, qs, agrupamento):
        trunc_fn = TruncMonth('ifood_criado_em') if agrupamento == 'mes' else TruncDate('ifood_criado_em')

        rows = (
            qs
            .annotate(periodo=trunc_fn)
            .values('periodo')
            .annotate(
                pedidos=Count('id'),
                receita=Sum('total_valor'),
                cancelados=Count('id', filter=Q(status='CANCELLED')),
            )
            .order_by('periodo')
        )

        result = []
        for row in rows:
            total = row['pedidos'] or 0
            receita = float(row['receita'] or 0)
            cancelados = row['cancelados'] or 0
            nao_cancelados = total - cancelados
            ticket = receita / nao_cancelados if nao_cancelados else 0

            p = row['periodo']
            if hasattr(p, 'date'):
                p = p.date()

            label = p.strftime('%b/%Y') if agrupamento == 'mes' else p.strftime('%d/%m/%Y')

            result.append({
                'periodo': str(p),
                'label': label,
                'pedidos': total,
                'receita': round(receita, 2),
                'cancelados': cancelados,
                'ticket_medio': round(ticket, 2),
            })

        return result

    # ──────────────────────────────────────────────────────────────────────────

    def _export_excel(self, dados):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        CARAMELO = 'C97A3A'
        CINZA    = 'F5F5F5'

        def hfont(): return Font(bold=True, color='FFFFFF', size=11)
        def hfill(): return PatternFill('solid', fgColor=CARAMELO)
        def center(): return Alignment(horizontal='center', vertical='center')
        def tfont(): return Font(bold=True, color='FFFFFF')

        wb = openpyxl.Workbook()

        # ── Sheet 1: Resumo ────────────────────────────────────────────────────
        ws1 = wb.active
        ws1.title = 'Resumo'

        ws1.merge_cells('A1:B1')
        t = ws1['A1']
        t.value = (
            f'Relatório iFood  —  '
            f'{dados["periodo"]["inicio"]} a {dados["periodo"]["fim"]}'
        )
        t.font = Font(bold=True, size=13, color=CARAMELO)
        t.alignment = center()
        ws1.row_dimensions[1].height = 28
        ws1.append([])

        r = dados['resumo']
        summary = [
            ('Total de Pedidos',    r['total_pedidos']),
            ('Receita Total (R$)',  r['receita_total']),
            ('Ticket Médio (R$)',   r['ticket_medio']),
            ('Cancelados',         r['cancelados']),
            ('Taxa de Cancelamento', f'{r["taxa_cancelamento"]}%'),
            ('Delivery',           r['delivery']),
            ('Retirada (Takeout)', r['takeout']),
        ]

        ws1.append(['Indicador', 'Valor'])
        hr = ws1.max_row
        for col in range(1, 3):
            c = ws1.cell(hr, col)
            c.font, c.fill, c.alignment = hfont(), hfill(), center()

        for i, (label, val) in enumerate(summary, 1):
            ws1.append([label, val])
            rn = ws1.max_row
            ws1.cell(rn, 1).alignment = Alignment(horizontal='left', vertical='center')
            ws1.cell(rn, 2).alignment = Alignment(horizontal='right', vertical='center')
            if i % 2 == 0:
                for col in range(1, 3):
                    ws1.cell(rn, col).fill = PatternFill('solid', fgColor=CINZA)

        ws1.column_dimensions['A'].width = 28
        ws1.column_dimensions['B'].width = 20

        # ── Sheet 2: Por Período ───────────────────────────────────────────────
        ws2 = wb.create_sheet('Por Período')
        agrup = 'Mês' if dados['agrupamento'] == 'mes' else 'Data'
        headers = [agrup, 'Pedidos', 'Receita (R$)', 'Cancelados', 'Ticket Médio (R$)']
        ws2.append(headers)
        hr2 = ws2.max_row
        for col in range(1, len(headers) + 1):
            c = ws2.cell(hr2, col)
            c.font, c.fill, c.alignment = hfont(), hfill(), center()

        for i, row in enumerate(dados['agrupado'], 1):
            ws2.append([row['label'], row['pedidos'], row['receita'], row['cancelados'], row['ticket_medio']])
            rn = ws2.max_row
            ws2.cell(rn, 3).number_format = '#,##0.00'
            ws2.cell(rn, 5).number_format = '#,##0.00'
            if i % 2 == 0:
                for col in range(1, len(headers) + 1):
                    ws2.cell(rn, col).fill = PatternFill('solid', fgColor=CINZA)

        if dados['agrupado']:
            tp = sum(x['pedidos']    for x in dados['agrupado'])
            tr = sum(x['receita']    for x in dados['agrupado'])
            tc = sum(x['cancelados'] for x in dados['agrupado'])
            nc = tp - tc
            tkt = tr / nc if nc else 0
            ws2.append(['TOTAL', tp, round(tr, 2), tc, round(tkt, 2)])
            rn = ws2.max_row
            for col in range(1, len(headers) + 1):
                c = ws2.cell(rn, col)
                c.font, c.fill, c.alignment = tfont(), hfill(), center()
            ws2.cell(rn, 3).number_format = '#,##0.00'
            ws2.cell(rn, 5).number_format = '#,##0.00'

        for w, col in zip([18, 12, 18, 14, 20], 'ABCDE'):
            ws2.column_dimensions[col].width = w
        ws2.auto_filter.ref = f'A1:E{ws2.max_row}'

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f'relatorio_ifood_{dados["periodo"]["inicio"]}_{dados["periodo"]["fim"]}.xlsx'
        response = HttpResponse(buf, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response

    # ──────────────────────────────────────────────────────────────────────────

    def _export_pdf(self, dados):
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle
            from reportlab.lib.units import cm
            from reportlab.lib.enums import TA_CENTER, TA_RIGHT
            from reportlab.platypus import (
                SimpleDocTemplate, Table, TableStyle,
                Paragraph, Spacer, HRFlowable,
            )
        except ImportError:
            return HttpResponse(
                'Dependência reportlab não instalada. Execute: pip install reportlab',
                status=500,
            )

        CARAMELO = colors.HexColor('#C97A3A')
        CINZA    = colors.HexColor('#F5F5F5')
        CINZA_BD = colors.HexColor('#E7E5E4')

        title_s  = ParagraphStyle('t',  fontName='Helvetica-Bold', fontSize=15, textColor=CARAMELO, alignment=TA_CENTER, spaceAfter=4)
        sub_s    = ParagraphStyle('s',  fontName='Helvetica',      fontSize=9,  textColor=colors.grey, alignment=TA_CENTER, spaceAfter=10)
        sec_s    = ParagraphStyle('sc', fontName='Helvetica-Bold', fontSize=11, textColor=CARAMELO, spaceBefore=14, spaceAfter=6)
        footer_s = ParagraphStyle('f',  fontName='Helvetica',      fontSize=7,  textColor=colors.grey, alignment=TA_RIGHT)

        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)

        story = []
        story.append(Paragraph('Arretado Doces — Relatório Consolidado iFood', title_s))
        agrup_txt = 'Mensal' if dados['agrupamento'] == 'mes' else 'Diário'
        story.append(Paragraph(
            f'Período: {dados["periodo"]["inicio"]} a {dados["periodo"]["fim"]} &nbsp;|&nbsp; Agrupamento: {agrup_txt}',
            sub_s,
        ))
        story.append(HRFlowable(width='100%', thickness=2, color=CARAMELO, spaceAfter=8))

        # Resumo
        story.append(Paragraph('Resumo do Período', sec_s))
        r = dados['resumo']
        resumo_rows = [
            ['Indicador', 'Valor'],
            ['Total de Pedidos',         str(r['total_pedidos'])],
            ['Receita Total',            f'R$ {r["receita_total"]:.2f}'],
            ['Ticket Médio',             f'R$ {r["ticket_medio"]:.2f}'],
            ['Pedidos Cancelados',       f'{r["cancelados"]} ({r["taxa_cancelamento"]}%)'],
            ['Delivery',                 str(r['delivery'])],
            ['Retirada (Takeout)',       str(r['takeout'])],
        ]
        t_resumo = Table(resumo_rows, colWidths=[9*cm, 6*cm])
        t_resumo.setStyle(TableStyle([
            ('BACKGROUND',   (0, 0), (-1, 0), CARAMELO),
            ('TEXTCOLOR',    (0, 0), (-1, 0), colors.white),
            ('FONTNAME',     (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0), (-1, 0), 10),
            ('FONTSIZE',     (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, CINZA]),
            ('GRID',         (0, 0), (-1, -1), 0.5, CINZA_BD),
            ('LEFTPADDING',  (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING',   (0, 0), (-1, -1), 5),
            ('BOTTOMPADDING',(0, 0), (-1, -1), 5),
            ('ALIGN',        (1, 0), (1, -1), 'RIGHT'),
        ]))
        story.append(t_resumo)

        # Detalhamento
        story.append(Paragraph('Detalhamento por Período', sec_s))
        agrup_col = 'Mês' if dados['agrupamento'] == 'mes' else 'Data'
        det_rows = [[agrup_col, 'Pedidos', 'Receita (R$)', 'Cancelados', 'Ticket Médio']]
        for row in dados['agrupado']:
            det_rows.append([
                row['label'],
                str(row['pedidos']),
                f'R$ {row["receita"]:.2f}',
                str(row['cancelados']),
                f'R$ {row["ticket_medio"]:.2f}',
            ])

        if dados['agrupado']:
            tp = sum(x['pedidos']    for x in dados['agrupado'])
            tr = sum(x['receita']    for x in dados['agrupado'])
            tc = sum(x['cancelados'] for x in dados['agrupado'])
            nc = tp - tc
            tkt = tr / nc if nc else 0
            det_rows.append(['TOTAL', str(tp), f'R$ {tr:.2f}', str(tc), f'R$ {tkt:.2f}'])

        last = len(det_rows) - 1
        t_det = Table(det_rows, colWidths=[3.5*cm, 2.5*cm, 4*cm, 2.5*cm, 4*cm])
        ts = [
            ('BACKGROUND',   (0, 0),    (-1, 0),    CARAMELO),
            ('TEXTCOLOR',    (0, 0),    (-1, 0),    colors.white),
            ('FONTNAME',     (0, 0),    (-1, 0),    'Helvetica-Bold'),
            ('FONTSIZE',     (0, 0),    (-1, 0),    9),
            ('FONTSIZE',     (0, 1),    (-1, -1),   8),
            ('ALIGN',        (0, 0),    (-1, -1),   'CENTER'),
            ('ROWBACKGROUNDS', (0, 1), (-1, last - 1), [colors.white, CINZA]),
            ('GRID',         (0, 0),    (-1, -1),   0.5, CINZA_BD),
            ('TOPPADDING',   (0, 0),    (-1, -1),   4),
            ('BOTTOMPADDING',(0, 0),    (-1, -1),   4),
        ]
        if len(det_rows) > 1:
            ts += [
                ('BACKGROUND', (0, last), (-1, last), CARAMELO),
                ('TEXTCOLOR',  (0, last), (-1, last), colors.white),
                ('FONTNAME',   (0, last), (-1, last), 'Helvetica-Bold'),
            ]
        t_det.setStyle(TableStyle(ts))
        story.append(t_det)

        story.append(Spacer(1, 0.5*cm))
        story.append(HRFlowable(width='100%', thickness=1, color=CINZA_BD))
        story.append(Paragraph(
            f'Gerado em {timezone.now().strftime("%d/%m/%Y às %H:%M")} — Arretado Doces CRM',
            footer_s,
        ))

        doc.build(story)
        buf.seek(0)

        fname = f'relatorio_ifood_{dados["periodo"]["inicio"]}_{dados["periodo"]["fim"]}.pdf'
        response = HttpResponse(buf, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{fname}"'
        return response


CANAIS_VALIDOS = ('ifood', 'pdv', 'eventos')


class ProdutosMaisVendidosView(CsrfExemptMixin, views.APIView):
    """
    Ranking de produtos mais vendidos, consolidando iFood + PDV + Eventos.

    Só considera pedido/evento que representa venda de fato concretizada:
    iFood status=CONCLUDED, PDV status confirmado/em_preparo/pronto/concluido
    (exclui aberto/cancelado), Evento status=entregue. Orçamentos ficam de
    fora de propósito — são cotação, não venda fechada.
    """
    permission_classes = [AllowAny]

    def get(self, request):
        params = request.query_params
        hoje = timezone.localtime(timezone.now()).date()
        try:
            data_inicio = date.fromisoformat(params['data_inicio']) if params.get('data_inicio') else hoje - timedelta(days=29)
        except ValueError:
            data_inicio = hoje - timedelta(days=29)
        try:
            data_fim = date.fromisoformat(params['data_fim']) if params.get('data_fim') else hoje
        except ValueError:
            data_fim = hoje
        if data_inicio > data_fim:
            data_inicio, data_fim = data_fim, data_inicio

        canais = [c for c in params.getlist('canal') if c in CANAIS_VALIDOS] or list(CANAIS_VALIDOS)

        ordenar = params.get('ordenar', 'quantidade')
        if ordenar not in ('quantidade', 'valor'):
            ordenar = 'quantidade'

        try:
            limit = int(params.get('limit', 30))
        except (TypeError, ValueError):
            limit = 30
        limit = max(1, min(limit, 200))

        agregados = {}
        if 'ifood' in canais:
            self._somar(agregados, 'ifood', self._qs_ifood(data_inicio, data_fim))
        if 'pdv' in canais:
            self._somar(agregados, 'pdv', self._qs_pdv(data_inicio, data_fim))
        if 'eventos' in canais:
            self._somar(agregados, 'eventos', self._qs_eventos(data_inicio, data_fim))

        produtos = []
        for item in agregados.values():
            quantidade_total = sum(c['quantidade'] for c in item['canais'].values())
            valor_total = sum(c['valor'] for c in item['canais'].values())
            produtos.append({
                'nome': item['nome'],
                'quantidade_total': quantidade_total,
                'valor_total': round(valor_total, 2),
                'canais': {
                    canal: {'quantidade': v['quantidade'], 'valor': round(v['valor'], 2)}
                    for canal, v in item['canais'].items()
                },
            })

        produtos.sort(
            key=lambda p: p['quantidade_total'] if ordenar == 'quantidade' else p['valor_total'],
            reverse=True,
        )

        resumo = {
            'produtos_distintos': len(produtos),
            'quantidade_total': sum(p['quantidade_total'] for p in produtos),
            'valor_total': round(sum(p['valor_total'] for p in produtos), 2),
        }

        return Response({
            'periodo': {'inicio': str(data_inicio), 'fim': str(data_fim)},
            'canais': canais,
            'ordenar': ordenar,
            'resumo': resumo,
            'produtos': produtos[:limit],
        })

    # ── Querysets por canal ──────────────────────────────────────────────

    def _qs_ifood(self, data_inicio, data_fim):
        return (
            ItemPedidoIFood.objects
            .filter(
                pedido__status='CONCLUDED',
                pedido__ifood_criado_em__date__gte=data_inicio,
                pedido__ifood_criado_em__date__lte=data_fim,
            )
            .values('nome')
            .annotate(quantidade=Sum('quantidade'), valor=Sum('preco_total'))
        )

    def _qs_pdv(self, data_inicio, data_fim):
        return (
            ItemPedidoPDV.objects
            .filter(
                pedido__status__in=['confirmado', 'em_preparo', 'pronto', 'concluido'],
                pedido__criado_em__date__gte=data_inicio,
                pedido__criado_em__date__lte=data_fim,
            )
            .values('nome')
            .annotate(quantidade=Sum('quantidade'), valor=Sum('preco_total'))
        )

    def _qs_eventos(self, data_inicio, data_fim):
        return (
            ItemEvento.objects
            .filter(
                evento__status='entregue',
                evento__data_evento__gte=data_inicio,
                evento__data_evento__lte=data_fim,
            )
            .values('nome')
            .annotate(quantidade=Sum('quantidade'), valor=Sum('preco_total'))
        )

    # ── Merge cross-canal por nome normalizado ────────────────────────────

    def _somar(self, agregados, canal, qs):
        for row in qs:
            nome = row['nome'] or '(sem nome)'
            chave = _normalizar_nome(nome)
            if chave not in agregados:
                agregados[chave] = {'nome': nome, 'canais': {}}
            bucket = agregados[chave]['canais'].setdefault(canal, {'quantidade': 0, 'valor': 0.0})
            bucket['quantidade'] += row['quantidade'] or 0
            bucket['valor'] += float(row['valor'] or 0)
