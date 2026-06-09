"""
Management command: importar_planilha
Lê PLANILHA_DE_PRECIFICAÇÃO__ARRETADO.xlsx e popula:
  - fichas.MateriaPrima     (aba MATERIA PRIMA)
  - pdv.Produto             (aba Cálculo)
  - fichas.FichaTecnica     (aba PRODUTOS)
  - fichas.ItemFichaTecnica (aba PRODUTOS)
"""
import os
import re
from decimal import Decimal, ROUND_HALF_UP

from django.core.management.base import BaseCommand, CommandError

try:
    import openpyxl
except ImportError:
    raise ImportError("Instale openpyxl: pip install openpyxl")


# ── helpers ───────────────────────────────────────────────────────────────────

def _normalizar(s):
    return str(s).strip().upper() if s else ''


def _parse_unidade(unidade_compra_str):
    """
    Retorna (quantidade_compra, unidade_medida) a partir de strings como
    '1kg', '500g', '200ml', '1l', '30 unidades', 'LATA'.
    """
    s = str(unidade_compra_str).lower().strip()
    # kg → gramas
    m = re.match(r'([\d.,]+)\s*kg', s)
    if m:
        return Decimal(m.group(1).replace(',', '.')) * 1000, 'g'
    # g
    m = re.match(r'([\d.,]+)\s*g\b', s)
    if m:
        return Decimal(m.group(1).replace(',', '.')), 'g'
    # litros
    m = re.match(r'([\d.,]+)\s*l\b', s)
    if m:
        return Decimal(m.group(1).replace(',', '.')) * 1000, 'ml'
    # ml
    m = re.match(r'([\d.,]+)\s*ml', s)
    if m:
        return Decimal(m.group(1).replace(',', '.')), 'ml'
    # unidades
    m = re.match(r'([\d.,]+)\s*(un|unidade)', s)
    if m:
        return Decimal(m.group(1).replace(',', '.')), 'un'
    # LATA / outros → assume 1 unidade
    return Decimal('1'), 'un'


def _segmento_por_preco(preco):
    if preco <= Decimal('2.10'):
        return 'unidade_pequena'
    if preco <= Decimal('4.50'):
        return 'unidade_media'
    if preco <= Decimal('6.00'):
        return 'bem_casado'
    return 'bolo_encomenda'


# ── grupos de colunas da aba PRODUTOS ─────────────────────────────────────────
# Cada grupo = (col_titulo, col_index, col_nome, col_qty, col_cost)  — 0-based
GRUPOS_COLS = [
    (1,  1,  2,  3,  4),
    (6,  6,  7,  8,  9),
    (12, 12, 13, 14, 15),
    (17, 17, 18, 19, 20),
]


def _cell(row, idx):
    try:
        return row[idx]
    except IndexError:
        return None


class Command(BaseCommand):
    help = 'Importa matérias-primas, produtos e fichas técnicas da planilha XLSX'

    def add_arguments(self, parser):
        parser.add_argument(
            '--arquivo', default='PLANILHA_DE_PRECIFICAÇÃO__ARRETADO.xlsx',
            help='Caminho para o arquivo .xlsx',
        )
        parser.add_argument('--dry-run',         action='store_true', help='Simula sem salvar')
        parser.add_argument('--apenas-materias', action='store_true', help='Importa só matérias-primas')
        parser.add_argument('--sobrescrever',    action='store_true', help='Atualiza registros existentes')

    def handle(self, *args, **options):
        arquivo      = options['arquivo']
        dry_run      = options['dry_run']
        so_materias  = options['apenas_materias']
        sobrescrever = options['sobrescrever']

        if not os.path.exists(arquivo):
            raise CommandError(f'Arquivo não encontrado: {arquivo}')

        self.stdout.write(f'Abrindo {arquivo}...')
        wb = openpyxl.load_workbook(arquivo, data_only=True)

        mp_count     = self._importar_materias(wb, dry_run, sobrescrever)
        prod_count   = 0
        ficha_count  = 0

        if not so_materias:
            prod_count  = self._importar_produtos(wb, dry_run, sobrescrever)
            ficha_count = self._importar_fichas(wb, dry_run)

        self.stdout.write(self.style.SUCCESS(
            f'\n✓ Matérias-primas: {mp_count} | Produtos: {prod_count} | Fichas: {ficha_count}'
            + (' [DRY-RUN]' if dry_run else '')
        ))

    # ── Matérias-primas ───────────────────────────────────────────────────────

    def _importar_materias(self, wb, dry_run, sobrescrever):
        from fichas.models import MateriaPrima

        ws = wb['MATERIA PRIMA']
        rows = list(ws.iter_rows(values_only=True))

        # Bloco esquerdo (B=1, C=2, D=3) começa linha 8 (índice 7)
        esq = {}
        for row in rows[7:]:
            nome_raw  = _cell(row, 1)
            unid_raw  = _cell(row, 2)
            valor_raw = _cell(row, 3)
            if nome_raw and valor_raw and str(nome_raw).strip().upper() not in ('PRODUTOS', 'EXEMPLO: TOMATE', 'EXEMPLO: LEITE'):
                nome = str(nome_raw).strip()
                esq[_normalizar(nome)] = (nome, str(unid_raw or ''), Decimal(str(valor_raw)))

        # Bloco direito (F=5, G=6, H=7) começa linha 6 (índice 5)
        dir_ = {}
        for row in rows[5:]:
            nome_raw = _cell(row, 5)
            qty_raw  = _cell(row, 6)
            if nome_raw and qty_raw:
                key = _normalizar(nome_raw)
                dir_[key] = Decimal(str(qty_raw))

        count = 0
        for key, (nome, unid_str, valor) in esq.items():
            quantidade, unidade = _parse_unidade(unid_str)

            # Preferir quantidade do bloco direito quando disponível
            if key in dir_:
                quantidade = dir_[key]
            elif unid_str.upper() == 'LATA':
                # Leite condensado: 395g (valor padrão)
                quantidade = Decimal('395')
                unidade    = 'g'

            if dry_run:
                self.stdout.write(f'  [DRY] MateriaPrima: {nome} | {unid_str} | qty={quantidade} {unidade} | R${valor}')
                count += 1
                continue

            if sobrescrever:
                obj, created = MateriaPrima.objects.update_or_create(
                    nome=nome,
                    defaults=dict(
                        unidade_compra=unid_str,
                        quantidade_compra=quantidade,
                        unidade_medida=unidade,
                        valor_compra=valor,
                    ),
                )
            else:
                obj, created = MateriaPrima.objects.get_or_create(
                    nome=nome,
                    defaults=dict(
                        unidade_compra=unid_str,
                        quantidade_compra=quantidade,
                        unidade_medida=unidade,
                        valor_compra=valor,
                    ),
                )
            if created:
                count += 1
                self.stdout.write(f'  + MateriaPrima: {nome}')
            elif sobrescrever:
                count += 1
                self.stdout.write(f'  ~ MateriaPrima: {nome}')

        return count

    # ── Produtos (aba Cálculo) ────────────────────────────────────────────────

    def _importar_produtos(self, wb, dry_run, sobrescrever):
        from pdv.models import Produto, CategoriaProduto

        ws   = wb['Cálculo']
        rows = list(ws.iter_rows(values_only=True))

        # Linhas de produto: índice=1 é número sequencial, 2=nome, 4=preço
        count = 0
        for row in rows[3:]:
            seq_raw   = _cell(row, 1)
            nome_raw  = _cell(row, 2)
            preco_raw = _cell(row, 4)
            if not isinstance(seq_raw, int) or not nome_raw or not preco_raw:
                continue
            if preco_raw == 243:   # linha de exemplo/template com preço 243
                continue

            nome  = str(nome_raw).strip().title()
            preco = Decimal(str(preco_raw)).quantize(Decimal('0.01'), rounding=ROUND_HALF_UP)
            seg   = _segmento_por_preco(preco)

            if dry_run:
                self.stdout.write(f'  [DRY] Produto: {nome} | R${preco} | {seg}')
                count += 1
                continue

            if sobrescrever:
                obj, created = Produto.objects.update_or_create(
                    nome=nome,
                    defaults=dict(preco=preco, segmento=seg, disponivel_eventos=True),
                )
            else:
                obj, created = Produto.objects.get_or_create(
                    nome=nome,
                    defaults=dict(preco=preco, segmento=seg, disponivel_eventos=True),
                )
            if created or sobrescrever:
                count += 1
                self.stdout.write(f'  {"+" if created else "~"} Produto: {nome} (R${preco})')

        return count

    # ── Fichas técnicas (aba PRODUTOS) ────────────────────────────────────────

    def _importar_fichas(self, wb, dry_run):
        from fichas.models import MateriaPrima, FichaTecnica, ItemFichaTecnica
        from pdv.models import Produto

        ws   = wb['PRODUTOS']
        rows = list(ws.iter_rows(values_only=True))

        # Cada "section" começa numa linha onde algum grupo tem um nome de produto
        # (string não vazia, não é número, não é 'TOTAL', '#', 'INGREDIENTES')
        SKIP = {'TOTAL', '#', 'INGREDIENTES', 'QUANTIDADE EM (G,ML)  ', 'CUSTO/PROPORCIONAL',
                'PRODUTOS', ''}

        fichas_data   = {}   # nome → {itens: [(mp_nome, qty)], embalagem: 0.08}
        fichas_ativas = {}   # (grupo_idx) → nome_da_ficha_corrente

        def _is_product_title(val):
            if not val or not isinstance(val, str):
                return False
            v = val.strip().upper()
            return v and v not in SKIP and not v.startswith('EXEMPLO')

        for row in rows:
            row_list = list(row)

            # Detecta linhas de título (nome do produto nos cols de título dos grupos)
            for g_title, g_idx, g_nome, g_qty, g_cost in GRUPOS_COLS:
                title_val = _cell(row_list, g_title)
                if _is_product_title(title_val):
                    nome_ficha = str(title_val).strip().title()
                    fichas_ativas[g_title] = nome_ficha
                    if nome_ficha not in fichas_data:
                        fichas_data[nome_ficha] = {'itens': [], 'embalagem': Decimal('0.08')}

            # Processa linhas de ingrediente nos grupos ativos
            for g_title, g_idx, g_nome, g_qty, g_cost in GRUPOS_COLS:
                ficha_nome = fichas_ativas.get(g_title)
                if not ficha_nome:
                    continue

                idx_val  = _cell(row_list, g_idx)
                nome_val = _cell(row_list, g_nome)
                qty_val  = _cell(row_list, g_qty)
                cost_val = _cell(row_list, g_cost)

                if not isinstance(idx_val, int):
                    # Verifica se é linha de TOTAL (encerra ficha)
                    if nome_val and str(nome_val).strip().upper() == 'TOTAL':
                        fichas_ativas.pop(g_title, None)
                    continue

                nome_ing = str(nome_val or '').strip()
                if not nome_ing:
                    continue

                if nome_ing.lower() == 'embalagem':
                    emb = cost_val or 0.08
                    fichas_data[ficha_nome]['embalagem'] = Decimal(str(emb))
                elif qty_val is not None and isinstance(qty_val, (int, float)):
                    fichas_data[ficha_nome]['itens'].append((nome_ing, Decimal(str(qty_val))))

        count = 0
        for nome_ficha, dados in fichas_data.items():
            if not dados['itens']:
                self.stdout.write(f'  ! Ficha sem ingredientes quantificados: {nome_ficha} — ignorada')
                continue

            # Tenta localizar produto PDV pelo nome
            try:
                produto = Produto.objects.get(nome__iexact=nome_ficha)
                produto_id = produto.id
            except Produto.DoesNotExist:
                # Busca parcial
                qs = Produto.objects.filter(nome__icontains=nome_ficha.split()[0])
                produto_id = qs.first().id if qs.exists() else None
            except Produto.MultipleObjectsReturned:
                produto_id = None

            if dry_run:
                self.stdout.write(f'  [DRY] Ficha: {nome_ficha} ({len(dados["itens"])} itens) → produto_id={produto_id}')
                count += 1
                continue

            ficha, created = FichaTecnica.objects.get_or_create(
                nome=nome_ficha,
                defaults=dict(
                    produto_pdv_id=produto_id,
                    embalagem_custo=dados['embalagem'],
                ),
            )
            if not created:
                self.stdout.write(f'  = Ficha já existe: {nome_ficha}')
                continue

            for mp_nome, qty in dados['itens']:
                # Busca matéria-prima pelo nome (case-insensitive, parcial)
                mp = (
                    MateriaPrima.objects.filter(nome__iexact=mp_nome).first()
                    or MateriaPrima.objects.filter(nome__icontains=mp_nome.split()[0]).first()
                )
                if mp:
                    ItemFichaTecnica.objects.get_or_create(
                        ficha=ficha, materia_prima=mp,
                        defaults={'quantidade': qty},
                    )
                else:
                    # Cria matéria-prima com custo zero para não travar a importação
                    mp, _ = MateriaPrima.objects.get_or_create(
                        nome=mp_nome,
                        defaults=dict(
                            unidade_compra='?',
                            quantidade_compra=Decimal('1'),
                            unidade_medida='un',
                            valor_compra=Decimal('0.01'),
                        ),
                    )
                    ItemFichaTecnica.objects.get_or_create(
                        ficha=ficha, materia_prima=mp,
                        defaults={'quantidade': qty},
                    )
                    self.stdout.write(
                        self.style.WARNING(f'    ⚠ MateriaPrima criada com custo zero: {mp_nome}')
                    )

            self.stdout.write(f'  + Ficha: {nome_ficha} ({len(dados["itens"])} itens)')
            count += 1

        return count
